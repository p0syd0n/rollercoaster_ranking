from openpyxl import load_workbook
from enum import Enum
import numpy as np
import sys
from datetime import time
import heapq

class PriorityQueue:
    def __init__(self):
        self._heap = []

    def add(self, item, priority):
        heapq.heappush(self._heap, (priority, item))

    def pop(self):
        return heapq.heappop(self._heap)[1]

    def peek(self, n=1):
        """Return the top n items (smallest priorities) without removing them."""
        return [item for (_, item) in sorted(self._heap)[:n]]

    def __len__(self):
        return len(self._heap)


construction = 1
type_ = 2
height = 3
speed = 4
length = 5
inversions = 6
drop = 7
duration = 8
gforce = 9
vertical_angle = 10
HOWMANY = 4
class Construction(Enum):
    STEEL = 1
    WOOD = 2

class Type(Enum):
    SIT_DOWN = 1
    INVERTED = 2
    STAND_UP = 3
    SUSPENDED = 4
    FLYING = 5
    STEEL = 6
    WING = 7
    WOOD  = 8

class HL(Enum):
    HIGH = 2
    LOW = 1

def time_(t: time) -> int:
    return t.hour * 3600 + t.minute * 60 + t.second
# Load the workbook
workbook = load_workbook(filename="dataset.xlsx")
sheet = workbook.active

optimal_values = {
    "construction": Construction.STEEL,
    "type": Type.SIT_DOWN,
    "height": HL.LOW,
    "speed": HL.HIGH,
    "length": HL.HIGH,
    "inversions": HL.HIGH,  
    "drop": HL.HIGH,
    "duration": HL.HIGH,
    "gforce": HL.HIGH,
    "vertical_angle": HL.HIGH
}

optimal_vector = {
    "construction": Construction.STEEL.value,
    "type": Type.SIT_DOWN.value,
    "height": None,
    "speed": None,
    "length": None,
    "inversions": None,  # Fixed typo
    "drop": None,
    "duration": None,
    "gforce": None,
    "vertical_angle": None
}
rows = sheet.iter_rows()
next(rows)
for row in rows:
    if row[height].value == None:
        break

    # Height
    print(f"height: {row[height].value}")
    if optimal_vector["height"] is None or \
       ((optimal_values["height"] == HL.HIGH and row[height].value > optimal_vector["height"]) or \
        (optimal_values["height"] == HL.LOW and row[height].value < optimal_vector["height"])):
        optimal_vector["height"] = row[height].value
    
    # Speed
    if optimal_vector["speed"] is None or \
       ((optimal_values["speed"] == HL.HIGH and row[speed].value > optimal_vector["speed"]) or \
        (optimal_values["speed"] == HL.LOW and row[speed].value < optimal_vector["speed"])):
        optimal_vector["speed"] = row[speed].value
    
    # Length
    if optimal_vector["length"] is None or \
       ((optimal_values["length"] == HL.HIGH and row[length].value > optimal_vector["length"]) or \
        (optimal_values["length"] == HL.LOW and row[length].value < optimal_vector["length"])):
        optimal_vector["length"] = row[length].value
    
    # Inversions
    if optimal_vector["inversions"] is None or \
       ((optimal_values["inversions"] == HL.HIGH and row[inversions].value > optimal_vector["inversions"]) or \
        (optimal_values["inversions"] == HL.LOW and row[inversions].value < optimal_vector["inversions"])):
        optimal_vector["inversions"] = row[inversions].value
    
    # Drop
    if optimal_vector["drop"] is None or \
       ((optimal_values["drop"] == HL.HIGH and row[drop].value > optimal_vector["drop"]) or \
        (optimal_values["drop"] == HL.LOW and row[drop].value < optimal_vector["drop"])):
        optimal_vector["drop"] = row[drop].value
    
    # Duration
    if optimal_vector["duration"] is None or \
       ((optimal_values["duration"] == HL.HIGH and row[duration].value > optimal_vector["duration"]) or \
        (optimal_values["duration"] == HL.LOW and row[duration].value < optimal_vector["duration"])):
        optimal_vector["duration"] = row[duration].value
    
    # G-Force
    if optimal_vector["gforce"] is None or \
       ((optimal_values["gforce"] == HL.HIGH and row[gforce].value > optimal_vector["gforce"]) or \
        (optimal_values["gforce"] == HL.LOW and row[gforce].value < optimal_vector["gforce"])):
        optimal_vector["gforce"] = row[gforce].value
    
    # Vertical Angle
    if optimal_vector["vertical_angle"] is None or \
       ((optimal_values["vertical_angle"] == HL.HIGH and row[vertical_angle].value > optimal_vector["vertical_angle"]) or \
        (optimal_values["vertical_angle"] == HL.LOW and row[vertical_angle].value < optimal_vector["vertical_angle"])):
        optimal_vector["vertical_angle"] = row[vertical_angle].value
print(f"Optimal vector: {optimal_vector}")
optimal_vector["duration"] = time_(optimal_vector["duration"])
rows = sheet.iter_rows()
next(rows)
thequeue = PriorityQueue()
optimal_vector = np.array(list(optimal_vector.values()))


coaster_to_name = {}
for row in rows:
    if row[height].value == None:
        break
    this_coaster = []
    match str(row[construction].value).strip():
        case "Steel":
            this_coaster.append(Construction.STEEL.value)
        case "Wood":
            this_coaster.append(Construction.WOOD.value)
        case _:
            print("Construction error")
            print(row[construction].value)
            print(row[construction])
            sys.exit()

    match str(row[type_].value).strip():
        case 'Sit Down':
            this_coaster.append(Type.SIT_DOWN.value)
        case 'Inverted':
            this_coaster.append(Type.INVERTED.value)
        case 'Stand Up':
            this_coaster.append(Type.STAND_UP.value)
        case 'Suspended':
            this_coaster.append(Type.SUSPENDED.value)
        case 'Flying':
            this_coaster.append(Type.FLYING.value)
        case 'Wood':
            this_coaster.append(Type.WOOD.value)
        case 'Steel':
            this_coaster.append(Type.STEEL.value)
        case 'Wing':
            this_coaster.append(Type.WING.value)
        case _:
            print("Type error")
            print(row[type_].value)
            print(row[type_])
            sys.exit()

    this_coaster.append(row[height].value)
    this_coaster.append(row[speed].value)
    this_coaster.append(row[length].value)
    this_coaster.append(row[inversions].value)
    this_coaster.append(row[drop].value)
    this_coaster.append(time_(row[duration].value))
    this_coaster.append(row[gforce].value)
    this_coaster.append(row[vertical_angle].value)
    this_coaster_vector = np.array(this_coaster)
    coaster_to_name[tuple(this_coaster)] = row[0].value
    opt = optimal_vector / np.linalg.norm(optimal_vector)
    vec = this_coaster_vector / np.linalg.norm(this_coaster_vector)
    score = -np.dot(opt, vec)

    thequeue.add(this_coaster, score)
    print(f"Added value. New length: {len(thequeue)}")
print("giving peek")
print(thequeue.peek(4))
for coaster in thequeue.peek(HOWMANY):
    print(f"Name: {coaster_to_name[tuple(coaster)]}\nData:{coaster}")
